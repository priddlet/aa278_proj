"""Monte Carlo campaign tests."""

import numpy as np
import pytest

from pulsar_nav.simulation.monte_carlo import (
    LUNANET_REQUIREMENT_M,
    MonteCarloConfig,
    aggregate_policy_stats,
    run_monte_carlo,
    run_preset_comparison,
    run_toa_noise_sweep,
    select_pulsars,
    steady_state_arc_metrics,
)
from pulsar_nav.simulation.predict_mode import PredictMode
from pulsar_nav.simulation.policy import NavPolicy, PolicySegment
from pulsar_nav.spice.kernels import resolve_kernel_dir

spiceypy = pytest.importorskip("spiceypy")


def _kernels_available() -> bool:
    try:
        resolve_kernel_dir()
        return True
    except FileNotFoundError:
        return False


@pytest.fixture(scope="module")
def spice_loaded():
    from pulsar_nav.spice.kernels import load_kernels

    load_kernels(load_gps_frames=True)


def test_preset_comparison_preserves_predict_mode(monkeypatch):
    captured: list[MonteCarloConfig] = []

    def fake_run(cfg: MonteCarloConfig, *, propagate_once: bool = True):
        captured.append(cfg)
        from pulsar_nav.simulation.monte_carlo import MonteCarloResult

        return MonteCarloResult(config=cfg, trials=[], by_policy={})

    monkeypatch.setattr(
        "pulsar_nav.simulation.monte_carlo.run_monte_carlo",
        fake_run,
    )
    base = MonteCarloConfig(
        n_trials=1,
        predict_mode=PredictMode.DYNAMICS,
        use_truth_velocity_predict=False,
        use_dynamics_predict=True,
        include_disturbances=True,
    )
    run_preset_comparison(("elfo", "llo"), base)
    assert len(captured) == 2
    assert all(c.predict_mode == PredictMode.DYNAMICS for c in captured)
    assert all(c.include_disturbances for c in captured)


def test_toa_sweep_preserves_predict_mode(monkeypatch):
    """Sweep configs must inherit predict_mode from base (not default truth_vel)."""
    captured: list[MonteCarloConfig] = []

    def fake_run(cfg: MonteCarloConfig, *, propagate_once: bool = True):
        captured.append(cfg)
        from pulsar_nav.simulation.monte_carlo import MonteCarloResult

        return MonteCarloResult(config=cfg, trials=[], by_policy={})

    monkeypatch.setattr(
        "pulsar_nav.simulation.monte_carlo.run_monte_carlo",
        fake_run,
    )
    base = MonteCarloConfig(
        n_trials=1,
        predict_mode=PredictMode.CV,
        use_truth_velocity_predict=False,
        use_dynamics_predict=False,
    )
    run_toa_noise_sweep((1.0,), base_config=base)
    assert len(captured) == 1
    assert captured[0].predict_mode == PredictMode.CV
    assert captured[0].use_truth_velocity_predict is False


def test_steady_state_arc_metrics_last_ten_percent():
    """Tail uses last 10% of epochs; epoch-0 spike excluded for n=100."""
    errs = np.zeros(100)
    errs[0] = 1000.0
    errs[1:] = 0.1
    mean_m, rms_m = steady_state_arc_metrics(errs)
    assert mean_m == pytest.approx(0.1)
    assert rms_m == pytest.approx(0.1)
    nan_mean, nan_rms = steady_state_arc_metrics(np.array([]))
    assert np.isnan(nan_mean) and np.isnan(nan_rms)


def test_select_pulsars_subset():
    p1 = select_pulsars(1)
    p3 = select_pulsars(3)
    assert len(p1) == 1
    assert len(p3) == 3


def test_aggregate_policy_stats():
    from pulsar_nav.simulation.monte_carlo import TrialMetrics

    trials = [
        TrialMetrics(
            trial_id=0,
            policy=NavPolicy.HYBRID,
            final_error_m=10.0,
            mean_error_m=8.0,
            rms_error_m=9.0,
            p95_error_m=12.0,
            max_error_m=15.0,
            blackout_mean_m=11.0,
            non_blackout_mean_m=5.0,
            n_pulsars=5,
            toa_sigma_s=1e-4,
            position_offset_m=50e3,
            steady_state_mean_m=1.0,
            steady_state_rms_m=1.5,
        ),
        TrialMetrics(
            trial_id=1,
            policy=NavPolicy.HYBRID,
            final_error_m=20.0,
            mean_error_m=18.0,
            rms_error_m=19.0,
            p95_error_m=22.0,
            max_error_m=25.0,
            blackout_mean_m=21.0,
            non_blackout_mean_m=9.0,
            n_pulsars=5,
            toa_sigma_s=1e-4,
            position_offset_m=60e3,
            steady_state_mean_m=4.0,
            steady_state_rms_m=4.5,
        ),
    ]
    stats = aggregate_policy_stats(trials, NavPolicy.HYBRID)
    assert stats.final_mean_m == 15.0
    assert stats.steady_state_mean_m == 2.5
    assert stats.steady_state_rms_m == 3.0
    assert stats.n_trials == 2
    assert stats.meets_lunanet_p95 is False


def test_planned_and_measured_segments():
    from pulsar_nav.simulation.policy import planned_segment, segment_from_measurements
    from pulsar_nav.visibility.blackout import NavMode, VisibilitySample

    blk = VisibilitySample(
        t_s=0.0,
        earth_elevation_deg=-10.0,
        gnss_visible=False,
        lonet_visible=True,
        n_lonet_visible=2,
        max_lonet_elevation_deg=30.0,
        nav_mode=NavMode.LONET,
        in_blackout=True,
    )
    vis = VisibilitySample(
        t_s=1.0,
        earth_elevation_deg=20.0,
        gnss_visible=True,
        lonet_visible=True,
        n_lonet_visible=2,
        max_lonet_elevation_deg=30.0,
        nav_mode=NavMode.HYBRID,
        in_blackout=False,
    )
    assert planned_segment(NavPolicy.HYBRID, blk) == PolicySegment.XNAV_LONET_SUPPLEMENT
    assert segment_from_measurements(
        NavPolicy.HYBRID, blk, n_gnss=0, n_lonet=2, n_pulsar=5
    ) == PolicySegment.XNAV_LONET_SUPPLEMENT
    assert planned_segment(NavPolicy.GNSS_ONLY, blk) == PolicySegment.XNAV_LONET_SUPPLEMENT
    assert segment_from_measurements(
        NavPolicy.GNSS_ONLY, vis, n_gnss=0, n_lonet=0, n_pulsar=5
    ) == PolicySegment.GNSS_XNAV_FALLBACK
    assert planned_segment(NavPolicy.HYBRID, vis) == PolicySegment.HYBRID_VISIBLE


@pytest.mark.skipif(not _kernels_available(), reason="SPICE kernels not on disk")
def test_monte_carlo_runs_all_policies(spice_loaded):
    cfg = MonteCarloConfig(
        n_trials=2,
        seed=99,
        duration_s=1800.0,
        step_s=300.0,
        randomize_offset=False,
        position_offset_m=40_000.0,
        policies=(
            NavPolicy.XNAV_ONLY,
            NavPolicy.GNSS_ONLY,
            NavPolicy.HYBRID,
        ),
    )
    result = run_monte_carlo(cfg)
    assert len(result.trials) == 2 * 3
    assert NavPolicy.HYBRID in result.by_policy
    for pol in cfg.policies:
        assert result.by_policy[pol].n_trials == 2
        assert result.by_policy[pol].final_mean_m < 500_000.0


@pytest.mark.skipif(not _kernels_available(), reason="SPICE kernels not on disk")
def test_hybrid_with_broadcast_gps_campaign(spice_loaded):
    cfg = MonteCarloConfig(
        n_trials=2,
        seed=11,
        duration_s=6.0 * 3600.0,
        step_s=180.0,
        randomize_offset=False,
        position_offset_m=50_000.0,
        policies=(NavPolicy.HYBRID, NavPolicy.XNAV_ONLY),
    )
    result = run_monte_carlo(cfg)
    h = result.by_policy[NavPolicy.HYBRID]
    x = result.by_policy[NavPolicy.XNAV_ONLY]
    assert h.n_trials == 2
    assert h.final_mean_m < 100_000.0
    assert x.final_mean_m < 100_000.0


@pytest.mark.skipif(not _kernels_available(), reason="SPICE kernels not on disk")
def test_gnss_only_uses_pulsars_in_blackout(spice_loaded):
    from pulsar_nav.propagation.dynamics import DynamicsConfig
    from pulsar_nav.propagation.propagator import LunarPropagator
    from pulsar_nav.simulation.presentation_runs import run_representative_policy_runs
    from pulsar_nav.spice.ephemeris import str_to_et
    from pulsar_nav.visibility.blackout import compute_visibility_timeline

    cfg = MonteCarloConfig(
        n_trials=1,
        seed=3,
        preset="elfo_nav",
        duration_s=26.0 * 3600.0,
        step_s=180.0,
        randomize_offset=False,
        position_offset_m=50_000.0,
        policies=(NavPolicy.GNSS_ONLY, NavPolicy.GNSS_COAST),
    )
    et0 = str_to_et(cfg.epoch_utc)
    prop = LunarPropagator(et0, config=DynamicsConfig(), auto_load_kernels=False)
    traj = prop.propagate_preset(cfg.preset, duration_s=cfg.duration_s, step_s=cfg.step_s)
    timeline = compute_visibility_timeline(traj)
    runs, _ = run_representative_policy_runs(cfg, traj=traj, timeline=timeline)
    gnss_switch = runs[NavPolicy.GNSS_ONLY]
    coast = runs[NavPolicy.GNSS_COAST]
    blk = np.array([s.in_blackout for s in timeline.samples])
    assert np.any(blk)
    assert float(np.mean(gnss_switch.position_error_m[blk])) < float(
        np.mean(coast.position_error_m[blk])
    )


@pytest.mark.skipif(not _kernels_available(), reason="SPICE kernels not on disk")
def test_elfo_presets_differ_in_visibility(spice_loaded):
    from pulsar_nav.propagation.dynamics import DynamicsConfig
    from pulsar_nav.propagation.propagator import LunarPropagator
    from pulsar_nav.spice.ephemeris import str_to_et
    from pulsar_nav.visibility.blackout import compute_visibility_timeline

    et0 = str_to_et("2026-01-15 12:00:00")
    prop = LunarPropagator(et0, config=DynamicsConfig(), auto_load_kernels=False)
    traj_science = prop.propagate_preset("elfo", duration_s=6.0 * 3600.0, step_s=120.0)
    traj_nav = prop.propagate_preset("elfo_nav", duration_s=6.0 * 3600.0, step_s=120.0)
    tl_science = compute_visibility_timeline(traj_science)
    tl_nav = compute_visibility_timeline(traj_nav)

    assert abs(tl_science.blackout_fraction - tl_nav.blackout_fraction) > 0.25


@pytest.mark.skipif(not _kernels_available(), reason="SPICE kernels not on disk")
def test_monte_carlo_preset_comparison(spice_loaded):
    cfg = MonteCarloConfig(
        n_trials=2,
        seed=42,
        duration_s=6.0 * 3600.0,
        step_s=300.0,
        randomize_offset=False,
        position_offset_m=40_000.0,
        policies=(NavPolicy.HYBRID, NavPolicy.GNSS_COAST),
    )
    from pulsar_nav.simulation.monte_carlo import run_preset_comparison

    results = run_preset_comparison(("elfo", "elfo_nav"), cfg)
    assert set(results) == {"elfo", "elfo_nav"}
    assert (
        abs(
            results["elfo_nav"].timeline.blackout_fraction
            - results["elfo"].timeline.blackout_fraction
        )
        > 0.25
    )


@pytest.mark.skipif(not _kernels_available(), reason="SPICE kernels not on disk")
def test_export_monte_carlo_xlsx(spice_loaded, tmp_path):
    pytest.importorskip("openpyxl")
    from pulsar_nav.simulation.monte_carlo_export import (
        MonteCarloExportBundle,
        export_monte_carlo_xlsx,
    )

    cfg = MonteCarloConfig(
        n_trials=2,
        seed=1,
        duration_s=1800.0,
        step_s=300.0,
        randomize_offset=False,
        position_offset_m=40_000.0,
        policies=(NavPolicy.HYBRID, NavPolicy.XNAV_ONLY),
    )
    result = run_monte_carlo(cfg)
    path = export_monte_carlo_xlsx(
        tmp_path / "mc.xlsx",
        MonteCarloExportBundle(main=result),
    )
    assert path.exists()

    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert {"config", "summary", "trials"}.issubset(set(wb.sheetnames))
    assert wb["trials"].max_row >= 5


@pytest.mark.skipif(not _kernels_available(), reason="SPICE kernels not on disk")
def test_hybrid_beats_gnss_coast_in_blackout_heavy_arc(spice_loaded):
    cfg = MonteCarloConfig(
        n_trials=3,
        seed=7,
        preset="elfo_nav",
        duration_s=4.0 * 3600.0,
        step_s=180.0,
        randomize_offset=True,
        policies=(NavPolicy.HYBRID, NavPolicy.GNSS_COAST),
    )
    result = run_monte_carlo(cfg)
    h = result.by_policy[NavPolicy.HYBRID]
    g = result.by_policy[NavPolicy.GNSS_COAST]
    assert h.blackout_mean_m < g.blackout_mean_m
