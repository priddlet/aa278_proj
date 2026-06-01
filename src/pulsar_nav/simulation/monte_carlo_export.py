"""Export Monte Carlo campaigns to Excel (.xlsx)."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from pulsar_nav.simulation.monte_carlo import (
    LUNANET_REQUIREMENT_M,
    MonteCarloConfig,
    MonteCarloResult,
    PolicyStats,
    TrialMetrics,
)


@dataclass
class MonteCarloExportBundle:
    """Optional collection of campaigns from one demo run."""

    main: MonteCarloResult | None = None
    comparison: dict[str, MonteCarloResult] | None = None
    pulsar_sweep: dict[int, MonteCarloResult] | None = None
    toa_sweep: dict[float, MonteCarloResult] | None = None


def _require_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "Excel export requires openpyxl. Install with: pip install openpyxl"
        ) from exc
    return Workbook, get_column_letter


def _config_rows(config: MonteCarloConfig, *, campaign: str, blackout_fraction: float | None) -> list[dict[str, Any]]:
    rows = [
        {"campaign": campaign, "parameter": "preset", "value": config.preset},
        {"campaign": campaign, "parameter": "epoch_utc", "value": config.epoch_utc},
        {"campaign": campaign, "parameter": "duration_hr", "value": config.duration_s / 3600.0},
        {"campaign": campaign, "parameter": "step_s", "value": config.step_s},
        {"campaign": campaign, "parameter": "n_trials", "value": config.n_trials},
        {"campaign": campaign, "parameter": "seed", "value": config.seed},
        {"campaign": campaign, "parameter": "toa_sigma_us", "value": config.toa_sigma_s * 1e6},
        {"campaign": campaign, "parameter": "gnss_sigma_m", "value": config.gnss_sigma_m},
        {"campaign": campaign, "parameter": "lonet_sigma_m", "value": config.lonet_sigma_m},
        {"campaign": campaign, "parameter": "n_pulsars", "value": config.n_pulsars if config.n_pulsars is not None else "all"},
        {"campaign": campaign, "parameter": "randomize_offset", "value": config.randomize_offset},
        {"campaign": campaign, "parameter": "offset_min_km", "value": config.offset_min_m / 1000.0},
        {"campaign": campaign, "parameter": "offset_max_km", "value": config.offset_max_m / 1000.0},
        {"campaign": campaign, "parameter": "policies", "value": ", ".join(p.value for p in config.policies)},
        {
            "campaign": campaign,
            "parameter": "use_truth_velocity_predict",
            "value": config.use_truth_velocity_predict,
        },
        {
            "campaign": campaign,
            "parameter": "process_noise_accel",
            "value": config.process_noise_accel,
        },
        {"campaign": campaign, "parameter": "lunanet_requirement_m", "value": LUNANET_REQUIREMENT_M},
    ]
    if blackout_fraction is not None:
        rows.append(
            {
                "campaign": campaign,
                "parameter": "blackout_fraction_pct",
                "value": round(100.0 * blackout_fraction, 2),
            }
        )
    return rows


def _trial_row(campaign: str, trial: TrialMetrics) -> dict[str, Any]:
    return {
        "campaign": campaign,
        "trial_id": trial.trial_id,
        "policy": trial.policy.value,
        "final_error_km": trial.final_error_m / 1000.0,
        "mean_error_km": trial.mean_error_m / 1000.0,
        "rms_error_km": trial.rms_error_m / 1000.0,
        "p95_error_km": trial.p95_error_m / 1000.0,
        "max_error_km": trial.max_error_m / 1000.0,
        "blackout_mean_km": trial.blackout_mean_m / 1000.0,
        "non_blackout_mean_km": trial.non_blackout_mean_m / 1000.0,
        "n_pulsars": trial.n_pulsars,
        "toa_sigma_us": trial.toa_sigma_s * 1e6,
        "position_offset_km": trial.position_offset_m / 1000.0,
        "sweep_label": trial.sweep_label,
    }


def _summary_row(campaign: str, stats: PolicyStats) -> dict[str, Any]:
    return {
        "campaign": campaign,
        "policy": stats.policy.value,
        "n_trials": stats.n_trials,
        "final_mean_km": stats.final_mean_m / 1000.0,
        "final_std_km": stats.final_std_m / 1000.0,
        "final_p95_km": stats.final_p95_m / 1000.0,
        "mean_error_km": stats.mean_error_m / 1000.0,
        "rms_error_km": stats.rms_error_m / 1000.0,
        "blackout_mean_km": stats.blackout_mean_m / 1000.0,
        "non_blackout_mean_km": stats.non_blackout_mean_m / 1000.0,
        "meets_lunanet_p95": stats.meets_lunanet_p95,
    }


def _collect_campaigns(bundle: MonteCarloExportBundle) -> list[tuple[str, MonteCarloResult]]:
    campaigns: list[tuple[str, MonteCarloResult]] = []
    if bundle.main is not None:
        campaigns.append((f"main_{bundle.main.config.preset}", bundle.main))
    if bundle.comparison:
        for preset, result in bundle.comparison.items():
            campaigns.append((f"compare_{preset}", result))
    if bundle.pulsar_sweep:
        for n, result in sorted(bundle.pulsar_sweep.items()):
            campaigns.append((f"pulsar_n{n}", result))
    if bundle.toa_sweep:
        for sigma_us, result in sorted(bundle.toa_sweep.items()):
            campaigns.append((f"toa_{sigma_us:g}us", result))
    return campaigns


def _autosize_columns(ws, get_column_letter) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def _write_table_sheet(ws, rows: list[dict[str, Any]], get_column_letter) -> None:
    if not rows:
        ws.append(["(empty)"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    _autosize_columns(ws, get_column_letter)


def export_monte_carlo_xlsx(path: str | Path, bundle: MonteCarloExportBundle) -> Path:
    """
    Write all Monte Carlo campaigns in ``bundle`` to a multi-sheet Excel workbook.

    Sheets: ``config``, ``summary``, ``trials``.
    """
    Workbook, get_column_letter = _require_openpyxl()
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    campaigns = _collect_campaigns(bundle)
    if not campaigns:
        raise ValueError("No Monte Carlo results to export")

    config_rows: list[dict[str, Any]] = []
    summary_rows: list[dict[str, Any]] = []
    trial_rows: list[dict[str, Any]] = []

    for campaign, result in campaigns:
        bf = result.timeline.blackout_fraction if result.timeline else None
        config_rows.extend(_config_rows(result.config, campaign=campaign, blackout_fraction=bf))
        for policy in result.config.policies:
            summary_rows.append(_summary_row(campaign, result.by_policy[policy]))
        for trial in result.trials:
            trial_rows.append(_trial_row(campaign, trial))

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets = [
        ("config", config_rows),
        ("summary", summary_rows),
        ("trials", trial_rows),
    ]
    for title, rows in sheets:
        ws = wb.create_sheet(title=title)
        _write_table_sheet(ws, rows, get_column_letter)

    wb.save(out)
    return out
